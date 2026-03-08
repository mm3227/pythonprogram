#core/control/funciones/exportar_horarios.py
import os
from tkinter import filedialog, messagebox


# ===============================
# PARA EXCEL
# ===============================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ===============================
# PARA PDF
# ===============================
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


HEADERS = [
    "Programa", "Grupo", "Modalidad",
    "Materia", "Salón", "Profesor",
    "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Horas"
]


# ==========================================================
# EXPORTAR A EXCEL (UNA HOJA POR PROGRAMA)
# ==========================================================
def exportar_excel(win):

    if not hasattr(win, "_filas_por_programa"):
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    nombre_ciclo = os.path.splitext(
        os.path.basename(win._db_ciclo_path)
    )[0]

    nombre_sugerido = f"{nombre_ciclo}_horario.xlsx"

    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=nombre_sugerido,
        filetypes=[("Archivo Excel", "*.xlsx")],
        title="Guardar archivo Excel"
    )

    if not ruta:
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 🎨 Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    filas_por_programa = win._filas_por_programa

    for programa, filas in filas_por_programa.items():

        ws = wb.create_sheet(title=programa[:31])

        # 🔵 Encabezados
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        row_excel = 2

        for fila in filas:

            datos = [
                fila[0].cget("text"),
                fila[1].cget("text"),
                fila[2].cget("text"),
                fila[3].get(),
                fila[4].get(),
                fila[5].get(),
            ]

            for i in range(6, 12):
                datos.append(fila[i].get())

            datos.append(fila[12].get())

            for col, valor in enumerate(datos, start=1):
                cell = ws.cell(row=row_excel, column=col)
                cell.value = valor
                cell.alignment = center_align

            row_excel += 1

        # 📏 Auto ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(ruta)

    messagebox.showinfo("Exportación completa", f"Archivo guardado en:\n{ruta}")


# ==========================================================
# EXPORTAR A PDF (SEPARADO POR PROGRAMA)
# ==========================================================
def exportar_pdf(win):
    try:
        if not hasattr(win, "_filas_por_programa"):
            messagebox.showerror("Error", "No hay datos para exportar.")
            return

        nombre_ciclo = os.path.splitext(os.path.basename(win._db_ciclo_path))[0]
        nombre_sugerido = f"{nombre_ciclo}_horario.pdf"
        ruta = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=nombre_sugerido,
            filetypes=[("Archivo PDF", "*.pdf")],
            title="Guardar archivo PDF"
        )
        if not ruta:
            return
    except Exception as e:
        import traceback
        error_detallado = traceback.format_exc()
        messagebox.showerror("Error al exportar PDF",
                             f"Ocurrió un error:\n{str(e)}\n\nRevisa el archivo error_log.txt para más detalles.")
        # Guardar traza en un archivo para depuración
        with open("error_log.txt", "a", encoding="utf-8") as f:
            f.write(f"Error en exportar_pdf:\n{error_detallado}\n")

    # Configuración de página
    page_width, page_height = landscape(letter)
    margin = 0.5 * inch
    available_width = page_width - 2 * margin

    doc = SimpleDocTemplate(ruta, pagesize=landscape(letter),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin)
    elements = []
    styles = getSampleStyleSheet()

    # Estilo para párrafos centrados
    style_centered = ParagraphStyle('Centered', parent=styles['Normal'],
                                    alignment=1,  # 1 = CENTER
                                    fontSize=12,
                                    spaceAfter=6)

    style_title = ParagraphStyle('Title', parent=styles['Normal'],
                                 alignment=1,
                                 fontSize=14,
                                 fontName='Helvetica-Bold',
                                 spaceAfter=12)

    style_subtitle = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                    alignment=1,
                                    fontSize=12,
                                    fontName='Helvetica-Bold',
                                    textColor=colors.HexColor("#1F4E78"),
                                    spaceAfter=6)

    # Estilo para celdas con texto centrado en dos líneas
    style_cell = ParagraphStyle('Cell', parent=styles['Normal'],
                                alignment=1,
                                fontSize=8,
                                leading=10,
                                fontName='Helvetica')

    filas_por_programa = win._filas_por_programa

    for programa, filas in filas_por_programa.items():
        # --- Encabezados institucionales ---
        elements.append(Paragraph("Universidad Autónoma de Zacatecas", style_title))
        elements.append(Paragraph("Unidad Académica de Ingeniería Eléctrica", style_subtitle))
        elements.append(Paragraph(f"Programa Académico: {programa}", style_centered))
        elements.append(Spacer(1, 0.2 * inch))

        # --- Definir nuevos encabezados sin la columna "Programa" ---
        headers = [
            "   Grupo   ", "     Modalidad     ", "Materia", "Salón", "Profesor",
            "L", "M", "X", "J", "V", "S", "   Horas   "
        ]

        # Construir datos para la tabla
        data = [headers]

        for fila in filas:
            # Obtener valores
            grupo = fila[1].cget("text")
            modalidad = fila[2].cget("text")
            materia = fila[3].get()
            salon_completo = fila[4].get().strip()
            profesor = fila[5].get()

            # Procesar salón en dos líneas si tiene dos partes
            if " " in salon_completo:
                edificio, salon = salon_completo.split(" ", 1)
                salon_para_celda = f"{edificio}<br/>{salon}"
            else:
                salon_para_celda = salon_completo

            # Crear Paragraph para el salón
            salon_paragraph = Paragraph(salon_para_celda, style_cell)

            # Horarios de días (L a S)
            horarios_dias = []
            for i in range(6, 12):
                horario = fila[i].get().strip()
                if horario and "-" in horario:
                    inicio, fin = horario.split("-")
                    # Mostrar inicio y fin en dos líneas
                    horario_para_celda = f"{inicio.strip()}<br/>{fin.strip()}"
                else:
                    horario_para_celda = ""
                horarios_dias.append(Paragraph(horario_para_celda, style_cell))

            # Horas totales
            horas_totales = fila[12].get()

            # Construir fila de datos
            fila_data = [
                Paragraph(grupo, style_cell),
                Paragraph(modalidad, style_cell),
                Paragraph(materia, style_cell),
                salon_paragraph,
                Paragraph(profesor, style_cell),
                horarios_dias[0],
                horarios_dias[1],
                horarios_dias[2],
                horarios_dias[3],
                horarios_dias[4],
                horarios_dias[5],
                Paragraph(horas_totales, style_cell)
            ]
            data.append(fila_data)

        # Calcular anchos máximos por columna (midiendo texto plano)
        num_cols = len(headers)
        max_widths = [stringWidth(headers[i], 'Helvetica-Bold', 8) for i in range(num_cols)]

        # Para las filas, necesitamos medir el contenido de los Paragraphs
        # Como los Paragraphs pueden tener múltiples líneas, usamos stringWidth sobre el texto plano
        for row in data[1:]:
            for i, cell in enumerate(row):
                if isinstance(cell, Paragraph):
                    # Extraer texto plano eliminando etiquetas HTML
                    text = cell.text.replace('<br/>', ' ').strip()
                else:
                    text = str(cell)
                w = stringWidth(text, 'Helvetica', 8)
                if w > max_widths[i]:
                    max_widths[i] = w

        # Ajustar al ancho disponible
        total = sum(max_widths)
        if total > available_width:
            factor = available_width / total
            col_widths = [w * factor for w in max_widths]
        else:
            col_widths = max_widths[:]

        # Redondear para evitar problemas
        col_widths = [round(w, 2) for w in col_widths]

        # Crear tabla
        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Estilo de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        elements.append(table)
        elements.append(PageBreak())
    
    # ---------------- IMAGEN DE ENCABEZADO ----------------   
    doc.build(elements)
    messagebox.showinfo("Exportación completa", f"PDF guardado en:\n{ruta}")