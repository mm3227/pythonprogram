import http.server
import socketserver
import socket
import sys
import urllib.parse
import tkinter as tk
from tkinter import messagebox
import sqlite3
import hashlib
import json
import secrets
import config

#==================================================================
from servidor_utils import obtener_ip_local, verificar_usuario, obtener_sesion, sesiones
from crear_db import crear_base_datos
from pgestor import profesores_api
from pgestor import programas_api
from pgestor import materias_api
from pgestor import salones_api
from pgestor import profesores_user_api
from pgestor import materias_user_api
from pgestor import grupos_api
#=============================================
from ciclos import ciclos_api

#==================================================================
#from para users
from pgestor import salones_user_api

PUERTO = 8000
#=======================================================   
class MiHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        ruta = self.path.split("?")[0]
        rol = obtener_sesion(self)
        # =====================================================
        # Rutas de API (no requieren protección de archivos)
        # =====================================================
        if self.path == "/listar_usuarios":
            conexion = sqlite3.connect("data/gestor.db")
            cursor = conexion.cursor()
            cursor.execute("SELECT id, usuario, programa, tipo_usuario FROM usuarios")
            filas = cursor.fetchall()
            conexion.close()
            datos = [{"id": f[0], "usuario": f[1], "programa": f[2], "tipo_usuario": f[3]} for f in filas]
            self.send_response(200)
            self.send_header("Content-type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(datos).encode())
            return

        elif self.path == "/listar_profesores":
            profesores_api.listar_profesores(self)
            return
        elif self.path == "/listar_programas":
            programas_api.listar_programas(self)
            return       
        elif self.path == "/listar_materias":
            materias_api.listar_materias(self)                  
        elif ruta == "/listar_grupos":
            grupos_api.listar_grupos(self)
            return
        elif self.path == "/listar_salones":
            salones_api.listar_salones(self)
            return
      

#===============================================================================
        #USERS LISTAR
        elif self.path == "/listar_salones_usuario":
            salones_user_api.listar_salones_usuario(self)
            return
        elif self.path == "/listar_profesores_usuario":
            profesores_user_api.listar_profesores_usuario(self)
            return
        elif self.path == "/listar_materias_usuario":
            materias_user_api.listar_materias_usuario(self)
            return
        
        
#===============================================================================
        elif self.path == "/listar_ciclos":
            ciclos_api.listar_ciclos(self)
            return
        """
        # =====================================================
        # Protección de páginas HTML
        # =====================================================
        # Protección de páginas ADMIN
        # ==========================================
        if self.path.startswith("/html/") or self.path == "/admin.html":

            if rol != "admin":

                self.send_response(302)
                self.send_header("Location","/index.html")
                self.end_headers()
                return


        # ==========================================
        # Protección de páginas USER
        # ==========================================
        if self.path.startswith("/htmlusers/") or self.path == "/users.html":

            if rol != "usuario":

                self.send_response(302)
                self.send_header("Location","/index.html")
                self.end_headers()
                return
        """
        # =====================================================
        # Para el resto de rutas, servir archivos estáticos
        # =====================================================
        super().do_GET()                                

    def do_POST(self):
        ruta = self.path.split("?")[0]
        # -------------------------
        # LOGIN
        # -------------------------
        if self.path == "/login":

            longitud = int(self.headers['Content-Length'])
            datos = self.rfile.read(longitud).decode()

            campos = urllib.parse.parse_qs(datos)

            usuario = campos.get("usuario", [""])[0]
            password = campos.get("password", [""])[0]

            rol = verificar_usuario(usuario, password)

            if rol:

                token = secrets.token_hex(16)
                sesiones[token] = {
                    "usuario": usuario,
                    "rol": rol
                }

                if rol == "admin":
                    destino = "/admin.html"
                else:
                    destino = "/users.html"

                self.send_response(302)

                self.send_header(
                    "Set-Cookie",
                    f"session={token}; HttpOnly"
                )

                self.send_header("Location", destino)
                self.end_headers()

            else:

                self.send_response(302)
                self.send_header("Location", "/index.html?error=1")
                self.end_headers()
                      
        # -------------------------
        # AGREGAR USUARIO
        # -------------------------
        elif self.path == "/agregar_usuario":

            longitud = int(self.headers['Content-Length'])
            datos = self.rfile.read(longitud).decode()

            datos_json = json.loads(datos)

            usuario = datos_json["usuario"]
            programa = datos_json["programa"]
            password = datos_json["password"]

            # SOLO cifrar contraseña
            hash_password = hashlib.sha256(password.encode()).hexdigest()

            conexion = sqlite3.connect("data/gestor.db")
            cursor = conexion.cursor()

            try:

                cursor.execute("""
                INSERT INTO usuarios (usuario, programa, password)
                VALUES (?, ?, ?)
                """, (usuario, programa, hash_password))

                conexion.commit()

                respuesta = "Usuario agregado correctamente"

            except sqlite3.IntegrityError:

                respuesta = "El usuario ya existe"

            conexion.close()

            self.send_response(200)
            self.end_headers()
            self.wfile.write(respuesta.encode())

        # -------------------------
        # IMPORTAR USUARIOS
        # -------------------------
        elif self.path == "/importar_usuarios":

            import cgi
            from openpyxl import load_workbook
            import io

            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    'REQUEST_METHOD': 'POST',
                    'CONTENT_TYPE': self.headers['Content-Type']
                }
            )

            archivo = form['archivo']

            if not archivo.file:

                self.send_response(400)
                self.end_headers()
                self.wfile.write(b"No se recibio archivo")
                return

            contenido = archivo.file.read()

            archivo_excel = io.BytesIO(contenido)

            wb = load_workbook(archivo_excel)
            hoja = wb.active

            conexion = sqlite3.connect("data/gestor.db")
            cursor = conexion.cursor()

            agregados = 0

            for fila in hoja.iter_rows(min_row=2, values_only=True):

                usuario, programa, password = fila

                if not usuario:
                    continue

                hash_password = hashlib.sha256(str(password).encode()).hexdigest()

                try:

                    cursor.execute("""
                    INSERT INTO usuarios (usuario, programa, password, tipo_usuario)
                    VALUES (?, ?, ?, ?)
                    """, (usuario, programa, hash_password, "usuario"))

                    agregados += 1

                except sqlite3.IntegrityError:
                    pass

            conexion.commit()
            conexion.close()

            self.send_response(200)
            self.end_headers()
            self.wfile.write(f"{agregados} usuarios importados".encode())
        # -------------------------
        # Reiniciar Contraseñas
        # -------------------------
        elif self.path == "/reset_password":

            longitud = int(self.headers['Content-Length'])
            datos = self.rfile.read(longitud).decode()

            datos_json = json.loads(datos)

            user_id = datos_json["id"]

            nueva_password = secrets.token_hex(4)

            hash_password = hashlib.sha256(nueva_password.encode()).hexdigest()

            conexion = sqlite3.connect("data/gestor.db")
            cursor = conexion.cursor()

            cursor.execute("""
            UPDATE usuarios
            SET password=?
            WHERE id=?
            """, (hash_password, user_id))

            conexion.commit()
            conexion.close()

            respuesta = {
                "password": nueva_password
            }

            self.send_response(200)
            self.send_header("Content-type","application/json")
            self.end_headers()

            self.wfile.write(json.dumps(respuesta).encode())
        
        # -------------------------
        # Eliminar usuario
        # -------------------------
        elif self.path == "/eliminar_usuario":

            longitud = int(self.headers['Content-Length'])
            datos = self.rfile.read(longitud).decode()

            datos_json = json.loads(datos)

            user_id = datos_json["id"]

            conexion = sqlite3.connect("data/gestor.db")
            cursor = conexion.cursor()

            # verificar tipo de usuario
            cursor.execute("SELECT tipo_usuario FROM usuarios WHERE id=?", (user_id,))
            tipo = cursor.fetchone()

            # si no existe el usuario
            if not tipo:
                self.send_response(200)
                self.end_headers()
                self.wfile.write(b"Usuario no encontrado")
                conexion.close()
                return

            # bloquear eliminación del admin
            if tipo[0] == "admin":

                self.send_response(200)
                self.end_headers()
                self.wfile.write(b"No se puede eliminar el administrador")
                conexion.close()
                return

            # eliminar usuario
            cursor.execute("DELETE FROM usuarios WHERE id=?", (user_id,))
            conexion.commit()
            conexion.close()

            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Usuario eliminado")
        
        #=================================================
        # Agregar, Eliminar, Editar e Importar  Profesores
        #=================================================
        elif self.path == "/agregar_profesor":
            profesores_api.agregar_profesor(self)
        elif self.path == "/eliminar_profesor":
            profesores_api.eliminar_profesor(self)
        elif self.path == "/editar_profesor":
            profesores_api.editar_profesor(self)
        elif self.path == "/importar_profesores":
            profesores_api.importar_profesores(self)
            
        #=================================================
        # Agregar, Eliminar, Editar e Importar  Materias
        #=================================================
        elif self.path == "/agregar_materia":
            materias_api.agregar_materia(self)
        elif self.path == "/eliminar_materia":
            materias_api.eliminar_materia(self)
        elif self.path == "/editar_materia":
            materias_api.editar_materia(self)          
        elif self.path == "/importar_materias":
            materias_api.importar_materias(self)
               
        #=================================================
        # Agregar, Eliminar, Editar e Importar  Salones
        #=================================================   
        elif self.path == "/agregar_salon":
            salones_api.agregar_salon(self)
        elif self.path == "/eliminar_salon":
            salones_api.eliminar_salon(self)
        elif self.path == "/editar_salon":
            salones_api.editar_salon(self)
        elif self.path == "/importar_salones":
            salones_api.importar_salones(self)
            
        #=================================================
        # users Salones
        #=================================================
        elif self.path == "/agregar_salon_usuario":
            salones_user_api.agregar_salon_usuario(self)
        elif self.path == "/eliminar_salon_usuario":
            salones_user_api.eliminar_salon_usuario(self)
        elif self.path == "/editar_salon_usuario":
            salones_user_api.editar_salon_usuario(self)
        elif self.path == "/importar_salones_usuario":
            salones_user_api.importar_salones_usuario(self)
        #=================================================
        # users Profesores
        #=================================================     
        elif self.path == "/agregar_profesor_usuario":
            profesores_user_api.agregar_profesor_usuario(self)
        elif self.path == "/eliminar_profesor_usuario":
            profesores_user_api.eliminar_profesor_usuario(self)
        elif self.path == "/editar_profesor_usuario":
            profesores_user_api.editar_profesor_usuario(self)
        elif self.path == "/importar_profesores_usuario":
            profesores_user_api.importar_profesores_usuario(self)
        #=================================================
        # users Materias
        #=================================================
        elif self.path == "/agregar_materia_usuario":
            materias_user_api.agregar_materia_usuario(self)
        elif self.path == "/eliminar_materia_usuario":
            materias_user_api.eliminar_materia_usuario(self)
        elif self.path == "/editar_materia_usuario":
            materias_user_api.editar_materia_usuario(self)
        elif self.path == "/importar_materias_usuario":
            materias_user_api.importar_materias_usuario(self)
        
        #=================================================
        # Agregar, Eliminar, Ciclos
        #=================================================
        elif self.path == "/crear_ciclo":
            ciclos_api.crear_ciclo(self)
        elif self.path == "/eliminar_ciclo":
            ciclos_api.eliminar_ciclo(self)
        #=================================================
        # Agregar, Eliminar, Editar e Importar Grupos
        #=================================================           
        elif ruta == "/agregar_grupo":
            grupos_api.agregar_grupo(self)
        elif ruta == "/eliminar_grupo":
            grupos_api.eliminar_grupo(self)
        elif ruta == "/editar_grupo":
            grupos_api.editar_grupo(self)
        elif ruta == "/importar_grupos":
            grupos_api.importar_grupos(self)
                               
        
#====================================================================================
def mostrar_info(ip_local, puerto, admin_creado, password):
    root = tk.Tk()
    root.withdraw()
    mensaje = f"Servidor iniciado\n\n"
    mensaje += f"Local:\nhttp://localhost:{puerto}\n\n"
    mensaje += f"Red:\nhttp://{ip_local}:{puerto}\n"
    if admin_creado:
        mensaje += "\nADMIN CREADO\n"
        mensaje += "usuario: admin\n"
        mensaje += f"password: {password}\n"
        mensaje += "\nGuarda esta contraseña."
    messagebox.showinfo("Servidor Web", mensaje)
    root.destroy()


def main():
    admin_creado, password = crear_base_datos()

    ip_local = obtener_ip_local()

    mostrar_info(ip_local, PUERTO, admin_creado, password)

    handler = MiHandler
    httpd = socketserver.ThreadingTCPServer(("", PUERTO), handler)

    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        httpd.shutdown()
        sys.exit(0)

if __name__ == "__main__":
    main()