import sqlite3
import json
import config
import cgi
import io
from openpyxl import load_workbook

def listar_profesores(handler):

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT id, programa, nombre, contratacion, telefono, email
    FROM profesores
    """)

    datos = cursor.fetchall()
    conexion.close()

    lista = []

    for p in datos:
        lista.append({
            "id": p[0],
            "programa": p[1],
            "nombre": p[2],
            "contratacion": p[3],
            "telefono": p[4],
            "email": p[5]
        })

    handler.send_response(200)
    handler.send_header("Content-type", "application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(lista).encode())
    
def agregar_profesor(handler):

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    programa = (datos_json["programa"] or "").strip()
    nombre = (datos_json["nombre"] or "").strip()
    contratacion = (datos_json["contratacion"] or "").strip()
    telefono = (datos_json["telefono"] or "").strip()
    email = (datos_json["email"] or "").strip()
    # validación mínima
    
    if not programa or not nombre:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"Programa y nombre son obligatorios")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    # =========================
    # verificar duplicado
    # =========================
    cursor.execute("""
    SELECT id FROM profesores
    WHERE nombre=? AND programa=?
    """, (nombre, programa))

    existe = cursor.fetchone()

    if existe:

        conexion.close()

        handler.send_response(409)
        handler.end_headers()
        handler.wfile.write(b"Profesor ya existe en este programa")
        return

    # =========================
    # insertar profesor
    # =========================
    cursor.execute("""
    INSERT INTO profesores
    (programa,nombre,contratacion,telefono,email)
    VALUES (?,?,?,?,?)
    """,(programa,nombre,contratacion,telefono,email))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Profesor agregado")
    
def eliminar_profesor(handler):

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    profesor_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("DELETE FROM profesores WHERE id=?", (profesor_id,))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Profesor eliminado")
    
def editar_profesor(handler):

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    programa = (datos_json["programa"] or "").strip()
    nombre = (datos_json["nombre"] or "").strip()
    contratacion = (datos_json["contratacion"] or "").strip()
    telefono = (datos_json["telefono"] or "").strip()
    email = (datos_json["email"] or "").strip()
    profesor_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()
    
    if not programa or not nombre:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"Programa y nombre son obligatorios")
        return
    
    cursor.execute("""
    UPDATE profesores
    SET programa=?, nombre=?, contratacion=?, telefono=?, email=?
    WHERE id=?
    """,(programa,nombre,contratacion,telefono,email,profesor_id))
    
    

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Profesor actualizado")
    
def importar_profesores(handler):

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': handler.headers['Content-Type']
        }
    )

    archivo = form['archivo']
    contenido = archivo.file.read()

    wb = load_workbook(io.BytesIO(contenido))
    hoja = wb.active

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    agregados = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        
        programa, nombre, contratacion, telefono, email = fila

        programa = str(programa or "").strip()
        nombre = str(nombre or "").strip()
        contratacion = str(contratacion or "").strip()
        telefono = str(telefono or "").strip()
        email = str(email or "").strip()

        if not programa or not nombre:
            continue

        try:

            cursor.execute("""
            INSERT INTO profesores
            (programa,nombre,contratacion,telefono,email)
            VALUES (?,?,?,?,?)
            """,(programa,nombre,contratacion,telefono,email))

            agregados += 1

        except sqlite3.IntegrityError:
            pass

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(f"{agregados} profesores importados".encode())