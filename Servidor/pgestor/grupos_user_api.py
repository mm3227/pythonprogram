from openpyxl import load_workbook
import cgi
import io
import sqlite3
import json
import config
import urllib.parse
from servidor_utils import obtener_programa_usuario

def listar_grupos_usuario(handler):

    programa = obtener_programa_usuario(handler)

    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sin programa")
        return

    # obtener parámetros
    query = urllib.parse.urlparse(handler.path).query
    params = urllib.parse.parse_qs(query)

    pagina = int(params.get("pagina", [1])[0])
    limite = int(params.get("limite", [10])[0])
    buscar = params.get("buscar", [""])[0]

    offset = (pagina - 1) * limite

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    # TOTAL (solo de su programa)
    cursor.execute("""
    SELECT COUNT(*) FROM grupos 
    WHERE programa = ? AND (
        grupo LIKE ? OR 
        semestre LIKE ? OR 
        tipo LIKE ?
    )
    """, (
        programa,
        f"%{buscar}%",
        f"%{buscar}%",
        f"%{buscar}%"
    ))
    total = cursor.fetchone()[0]

    # DATOS FILTRADOS
    cursor.execute("""
    SELECT id, semestre, grupo, tipo, alumnos, materias
    FROM grupos
    WHERE programa = ? AND (
        grupo LIKE ? OR 
        semestre LIKE ? OR 
        tipo LIKE ?
    )
    ORDER BY 
    CASE 
        WHEN tipo = 'Escolarizado' THEN 1
        WHEN tipo = 'Semiescolarizado' THEN 2
        ELSE 3
    END,
    semestre ASC,
    grupo COLLATE NOCASE ASC
    LIMIT ? OFFSET ?
    """, (
        programa,
        f"%{buscar}%",
        f"%{buscar}%",
        f"%{buscar}%",
        limite,
        offset
    ))

    datos = cursor.fetchall()
    conexion.close()

    lista = []

    for g in datos:
        lista.append({
            "id": g[0],
            "semestre": g[1],
            "grupo": g[2],
            "tipo": g[3],
            "alumnos": g[4],
            "materias": g[5]
        })

    respuesta = {
        "datos": lista,
        "total": total
    }

    handler.send_response(200)
    handler.send_header("Content-type","application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(respuesta).encode())

def agregar_grupo_usuario(handler):

    programa = obtener_programa_usuario(handler)
    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sin programa")
        return

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    semestre = datos_json["semestre"]
    grupo = datos_json["grupo"]
    tipo = datos_json["tipo"]
    alumnos = datos_json["alumnos"]
    materias = datos_json["materias"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    # 🔍 VALIDAR SI YA EXISTE
    cursor.execute("""
    SELECT 1 FROM grupos
    WHERE programa=? AND semestre=? AND grupo=? AND tipo=?
    """, (programa, semestre, grupo, tipo))

    if cursor.fetchone():
        conexion.close()
        handler.send_response(200)
        handler.end_headers()
        handler.wfile.write("⚠ Ese grupo ya existe en este programa".encode())
        return

    # INSERTAR
    cursor.execute("""
    INSERT INTO grupos
    (programa,semestre,grupo,tipo,alumnos,materias)
    VALUES (?,?,?,?,?,?)
    """,(programa,semestre,grupo,tipo,alumnos,materias))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Grupo agregado correctamente")
    
def eliminar_grupo_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    grupo_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    DELETE FROM grupos
    WHERE id=? AND programa=?
    """, (grupo_id, programa))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Grupo eliminado")
    
def importar_grupos_usuario(handler):  

    programa = obtener_programa_usuario(handler)
    #print("PROGRAMA IMPORT:", programa)
    #print("HEADERS:", handler.headers)

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD':'POST',
            'CONTENT_TYPE':handler.headers['Content-Type']
        }
    )

    archivo = form['archivo']
    contenido = archivo.file.read()

    wb = load_workbook(io.BytesIO(contenido))
    hoja = wb.active

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    agregados = 0
    repetidos = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        if not fila or len(fila) < 5:
            continue

        semestre, grupo, tipo, alumnos, materias = fila

        try:
            cursor.execute("""
            INSERT INTO grupos
            (programa,semestre,grupo,tipo,alumnos,materias)
            VALUES (?,?,?,?,?,?)
            """,(programa,semestre,grupo,tipo,alumnos,materias))

            agregados += 1

        except sqlite3.IntegrityError:
            repetidos += 1

    conexion.commit()
    conexion.close()

    mensaje = f"{agregados} grupos agregados, {repetidos} repetidos ignorados"

    handler.send_response(200)
    handler.send_header("Content-type", "text/plain")
    handler.end_headers()
    handler.wfile.write(mensaje.encode())
    print(programa, semestre, grupo, tipo)
    
    

