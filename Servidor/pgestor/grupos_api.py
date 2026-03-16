from openpyxl import load_workbook
import cgi
import io
import sqlite3
import json
import config
import urllib.parse

def listar_grupos(handler):

    import urllib.parse

    query = urllib.parse.urlparse(handler.path).query
    params = urllib.parse.parse_qs(query)

    pagina = int(params.get("pagina", [1])[0])
    limite = int(params.get("limite", [10])[0])
    buscar = params.get("buscar", [""])[0]  

    offset = (pagina - 1) * limite

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()
    
    #print("VALOR BUSCAR:", buscar)
    # TOTAL FILTRADO
    cursor.execute("""
    SELECT COUNT(*) FROM grupos
    WHERE programa LIKE ? OR grupo LIKE ? OR semestre LIKE ?
    """, (
        f"%{buscar}%",
        f"%{buscar}%",
        f"%{buscar}%"
    ))

    total = cursor.fetchone()[0]

    # DATOS FILTRADOS
    cursor.execute("""
    SELECT id, programa, semestre, grupo, tipo, alumnos, materias
    FROM grupos
    WHERE programa LIKE ? OR grupo LIKE ? OR semestre LIKE ?
    ORDER BY
    programa ASC,
    CASE 
        WHEN tipo = 'Escolarizado' THEN 1
        WHEN tipo = 'Semiescolarizado' THEN 2
        ELSE 3
    END,
    semestre ASC,
    grupo COLLATE NOCASE ASC
    LIMIT ? OFFSET ?
    """, (
        f"%{buscar}%",
        f"%{buscar}%",
        f"%{buscar}%",
        limite,
        offset
    ))

    datos = cursor.fetchall()
    conexion.close()

    lista = []

    for g in datos:
        lista.append({
            "id": g[0],
            "programa": g[1],
            "semestre": g[2],
            "grupo": g[3],
            "tipo": g[4],
            "alumnos": g[5],
            "materias": g[6]
        })

    respuesta = {
        "datos": lista,
        "total": total
    }

    handler.send_response(200)
    handler.send_header("Content-type", "application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(respuesta).encode())
    
def agregar_grupo(handler):
    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)
    programa = datos_json["programa"]
    semestre = datos_json["semestre"]
    grupo = datos_json["grupo"]
    tipo = datos_json["tipo"]
    alumnos = datos_json["alumnos"]
    materias =datos_json["materias"]
    
    conexion = sqlite3.connect(config.DB_PATH)

    cursor = conexion.cursor()

    try:

        cursor.execute("""
        INSERT INTO grupos
        (programa,semestre,grupo,tipo,alumnos,materias)
        VALUES (?,?,?,?,?,?)
        """,(programa,semestre,grupo,tipo,alumnos,materias))

        conexion.commit()

    except sqlite3.IntegrityError:

        handler.send_response(409)
        handler.end_headers()
        handler.wfile.write(b"El grupo ya existe")
        conexion.close()
        return

    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Grupo agregado")


def eliminar_grupo(handler):

    longitud = int(handler.headers['Content-Length'])

    datos = handler.rfile.read(longitud).decode()

    datos_json = json.loads(datos)

    grupo_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)

    cursor = conexion.cursor()

    cursor.execute("DELETE FROM grupos WHERE id=?", (grupo_id,))

    conexion.commit()

    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Grupo eliminado")
    
def importar_grupos(handler):

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD':'POST',
            'CONTENT_TYPE':handler.headers['Content-Type']
        }
    )

    archivo=form['archivo']

    contenido=archivo.file.read()

    archivo_excel=io.BytesIO(contenido)

    wb=load_workbook(archivo_excel)

    hoja=wb.active

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    agregados=0
    ignorados = 0

    for fila in hoja.iter_rows(min_row=2,values_only=True):

        programa,semestre,grupo,tipo,alumnos,materias=fila

        try:

            cursor.execute("""
            INSERT INTO grupos
            (programa,semestre,grupo,tipo,alumnos,materias)
            VALUES (?,?,?,?,?,?)
            """,(programa,semestre,grupo,tipo,alumnos,materias))

            agregados+=1

        except sqlite3.IntegrityError:
            ignorados += 1
            pass

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(f"{agregados} grupos importados, {ignorados} ignorados".encode())
    

def editar_grupo(handler):
    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    grupo_id = datos_json["id"]
    programa = datos_json["programa"]
    semestre = datos_json["semestre"]
    grupo = datos_json["grupo"]
    tipo = datos_json["tipo"]
    alumnos = datos_json["alumnos"]
    materias =datos_json["materias"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    try:
        cursor.execute("""
            UPDATE grupos
            SET programa=?, semestre=?, grupo=?, tipo=?, alumnos=?, materias=?
            WHERE id=?
        """, (programa, semestre, grupo, tipo, alumnos,materias, grupo_id))

        conexion.commit()
        respuesta = "Grupo actualizado"
    except sqlite3.IntegrityError:
        respuesta = "Error: el grupo ya existe"

    conexion.close()
    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(respuesta.encode())