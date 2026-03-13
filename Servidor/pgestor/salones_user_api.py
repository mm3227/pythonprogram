import sqlite3
import json
import config
import cgi
from openpyxl import load_workbook
import io
from servidor_utils import obtener_sesion, sesiones


def obtener_programa_usuario(handler):

    cookie = handler.headers.get("Cookie")

    if not cookie:
        return None

    token = None
    partes = cookie.split(";")

    for parte in partes:
        if "session=" in parte:
            token = parte.strip().split("=")[1]
            break

    if not token:
        return None

    sesion = sesiones.get(token)

    if not sesion:
        return None

    usuario = sesion["usuario"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT programa
    FROM usuarios
    WHERE usuario=?
    """,(usuario,))

    resultado = cursor.fetchone()

    conexion.close()

    if resultado:
        return resultado[0]

    return None


def listar_salones_usuario(handler):

    programa = obtener_programa_usuario(handler)
    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sesion invalida")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT id, edificio, salon, capacidad
    FROM salones
    WHERE programa=?
    """,(programa,))

    datos = cursor.fetchall()

    conexion.close()

    lista=[]

    for s in datos:

        lista.append({
            "id":s[0],
            "edificio":s[1],
            "salon":s[2],
            "capacidad":s[3]
        })

    handler.send_response(200)
    handler.send_header("Content-type","application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(lista).encode())
    
def agregar_salon_usuario(handler):
    programa = obtener_programa_usuario(handler)
    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()
    
    edificio = (datos_json["edificio"] or "").strip()
    salon = (datos_json["salon"] or "").strip()
    capacidad = datos_json["capacidad"] or 0

    cursor.execute("""
    INSERT INTO salones
    (programa,edificio,salon,capacidad)
    VALUES (?,?,?,?)
    """,(programa,edificio,salon,capacidad))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Salon agregado")
    
def editar_salon_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()

    datos_json = json.loads(datos)

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    UPDATE salones
    SET edificio=?, salon=?, capacidad=?
    WHERE id=? AND programa=?
    """,(

        datos_json["edificio"],
        datos_json["salon"],
        datos_json["capacidad"],
        datos_json["id"],
        programa

    ))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Salon actualizado")
    
def importar_salones_usuario(handler):

    programa_usuario = obtener_programa_usuario(handler)

    if not programa_usuario:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sesion invalida")
        return

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': handler.headers['Content-Type']
        }
    )

    if "archivo" not in form:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"No se recibio archivo")
        return

    archivo = form['archivo']

    if not archivo.file:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"No se recibio archivo")
        return

    contenido = archivo.file.read()
    archivo_excel = io.BytesIO(contenido)

    wb = load_workbook(archivo_excel)
    hoja = wb.active

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    agregados = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        if not fila:
            continue

        edificio = str(fila[0] or "").strip() if len(fila) > 0 else ""
        salon = str(fila[1] or "").strip() if len(fila) > 1 else ""
        capacidad = int(fila[2] or 0) if len(fila) > 2 else 0

        if not salon:
            continue

        try:

            cursor.execute("""
            INSERT INTO salones
            (programa, edificio, salon, capacidad)
            VALUES (?,?,?,?)
            """,(
                programa_usuario,
                edificio,
                salon,
                capacidad
            ))

            agregados += 1

        except sqlite3.IntegrityError:
            pass

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(f"{agregados} salones importados".encode())
    
def eliminar_salon_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()

    datos_json = json.loads(datos)

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    DELETE FROM salones
    WHERE id=? AND programa=?
    """,(

        datos_json["id"],
        programa

    ))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Salon eliminado")