import sqlite3
import json
import config
import cgi
from openpyxl import load_workbook
import io


def listar_materias(handler):
    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()
    cursor.execute("""
    SELECT id, programa, materia, continuidad, creditos
    FROM materias
    """)

    datos = cursor.fetchall()
    conexion.close()

    lista = []

    for m in datos:

        lista.append({
            "id":m[0],
            "programa":m[1],
            "materia":m[2],
            "continuidad":m[3],
            "creditos":m[4]
        })

    handler.send_response(200)
    handler.send_header("Content-type","application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(lista).encode())


def agregar_materia(handler):
    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    programa = (datos_json["programa"] or "").strip()
    materia = (datos_json["materia"] or "").strip()
    continuidad = (datos_json["continuidad"] or "").strip()
    creditos = datos_json["creditos"] or 0

    if not programa or not materia:

        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"Programa y materia son obligatorios")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT id FROM materias
    WHERE programa=? AND materia=?
    """,(programa,materia))

    existe = cursor.fetchone()

    if existe:
        conexion.close()
        handler.send_response(409)
        handler.end_headers()
        handler.wfile.write(b"La materia ya existe en este programa")
        return

    cursor.execute("""
    INSERT INTO materias
    (programa,materia,continuidad,creditos)
    VALUES (?,?,?,?)
    """,(programa,materia,continuidad,creditos))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Materia agregada")


def eliminar_materia(handler):
    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()

    datos_json = json.loads(datos)

    materia_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("DELETE FROM materias WHERE id=?", (materia_id,))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Materia eliminada")


def editar_materia(handler):

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    programa = (datos_json["programa"] or "").strip()
    materia = (datos_json["materia"] or "").strip()
    continuidad = (datos_json["continuidad"] or "").strip()
    creditos = datos_json["creditos"] or 0
    materia_id = datos_json["id"]

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    if not programa or not materia:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"Programa y materia son obligatorios")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    UPDATE materias
    SET programa=?, materia=?, continuidad=?, creditos=?
    WHERE id=?
    """,(programa,materia,continuidad,creditos,materia_id))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Materia actualizada")
    
def importar_materias(handler):
    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': handler.headers['Content-Type']
        }
    )

    archivo = form['archivo']

    contenido = archivo.file.read()

    archivo_excel = io.BytesIO(contenido)

    wb = load_workbook(archivo_excel)

    hoja = wb.active

    conexion = sqlite3.connect(config.DB_PATH)

    cursor = conexion.cursor()

    agregadas = 0

    for fila in hoja.iter_rows(min_row=2,values_only=True):

        programa,materia,continuidad,creditos = fila

        programa = (programa or "").strip()
        materia = (materia or "").strip()
        continuidad = (continuidad or "").strip()
        creditos = creditos or 0

        if not programa or not materia:
            continue

        try:

            cursor.execute("""
            INSERT INTO materias
            (programa,materia,continuidad,creditos)
            VALUES (?,?,?,?)
            """,(programa,materia,continuidad,creditos))

            agregadas +=1

        except sqlite3.IntegrityError:
            pass

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(f"{agregadas} materias importadas".encode())