import sqlite3
import json
import config
from servidor_utils import sesiones
import cgi
from openpyxl import load_workbook
import io


def obtener_programa_usuario(handler):

    cookie = handler.headers.get("Cookie")

    if not cookie:
        return None

    token=None

    partes=cookie.split(";")

    for parte in partes:

        if "session=" in parte:
            token=parte.strip().split("=")[1]
            break

    if not token:
        return None

    sesion=sesiones.get(token)

    if not sesion:
        return None

    usuario=sesion["usuario"]

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    SELECT programa
    FROM usuarios
    WHERE usuario=?
    """,(usuario,))

    resultado=cursor.fetchone()

    conexion.close()

    if resultado:
        return resultado[0]

    return None


# =========================================================
# LISTAR
# =========================================================

def listar_profesores_usuario(handler):

    programa = obtener_programa_usuario(handler)
    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sesion invalida")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT id,nombre,contratacion,telefono,email
    FROM profesores
    WHERE programa=?
    """,(programa,))

    datos = cursor.fetchall()

    conexion.close()

    lista=[]

    for p in datos:

        lista.append({

            "id":p[0],
            "nombre":p[1],
            "contratacion":p[2],
            "telefono":p[3],
            "email":p[4]

        })

    handler.send_response(200)
    handler.send_header("Content-type","application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(lista).encode())


# =========================================================
# AGREGAR
# =========================================================

def agregar_profesor_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()
    datos_json = json.loads(datos)

    nombre = (datos_json["nombre"] or "").strip()
    contratacion = (datos_json["contratacion"] or "").strip()
    telefono = (datos_json["telefono"] or "").strip()
    email = (datos_json["email"] or "").strip()

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    try:

        cursor.execute("""
        INSERT INTO profesores
        (programa,nombre,contratacion,telefono,email)
        VALUES (?,?,?,?,?)
        """,(programa,nombre,contratacion,telefono,email))

        conexion.commit()
        respuesta="Profesor agregado"

    except sqlite3.IntegrityError:
        respuesta="Profesor ya existe"

    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(respuesta.encode())


# =========================================================
# EDITAR
# =========================================================

def editar_profesor_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud=int(handler.headers['Content-Length'])
    datos=handler.rfile.read(longitud).decode()
    datos_json=json.loads(datos)

    nombre = (datos_json["nombre"] or "").strip()
    contratacion = (datos_json["contratacion"] or "").strip()
    telefono = (datos_json["telefono"] or "").strip()
    email = (datos_json["email"] or "").strip()

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    UPDATE profesores
    SET nombre=?,contratacion=?,telefono=?,email=?
    WHERE id=? AND programa=?
    """,(nombre,contratacion,telefono,email,datos_json["id"],programa))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()
    handler.wfile.write(b"Profesor actualizado")
    
# =========================================================
# ELIMINAR
# =========================================================

def eliminar_profesor_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud=int(handler.headers['Content-Length'])
    datos=handler.rfile.read(longitud).decode()

    datos_json=json.loads(datos)

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    DELETE FROM profesores
    WHERE id=? AND programa=?
    """,(

        datos_json["id"],
        programa

    ))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Profesor eliminado")


# =========================================================
# IMPORTAR EXCEL
# =========================================================

def importar_profesores_usuario(handler):

    programa = obtener_programa_usuario(handler)

    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sesion invalida")
        return

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': handler.headers['Content-Type']
        }
    )

    if "archivo" not in form:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"No se recibio archivo")
        return

    archivo = form["archivo"]

    if not archivo.file:
        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"No se recibio archivo")
        return

    contenido = archivo.file.read()

    archivo_excel = io.BytesIO(contenido)

    wb = load_workbook(archivo_excel)
    hoja = wb.active

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    agregados = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        if not fila:
            continue

        nombre = str(fila[0] or "").strip() if len(fila) > 0 else ""
        contratacion = str(fila[1] or "").strip() if len(fila) > 1 else ""
        telefono = str(fila[2] or "").strip() if len(fila) > 2 else ""
        email = str(fila[3] or "").strip() if len(fila) > 3 else ""

        if not nombre:
            continue

        try:

            cursor.execute("""
            INSERT INTO profesores
            (programa,nombre,contratacion,telefono,email)
            VALUES (?,?,?,?,?)
            """,(
                programa,
                nombre,
                contratacion,
                telefono,
                email
            ))

            agregados += 1

        except sqlite3.IntegrityError:
            pass

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(f"{agregados} profesores importados".encode())
