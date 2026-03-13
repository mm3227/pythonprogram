import sqlite3
import json
import config
from servidor_utils import sesiones

def obtener_programa_usuario(handler):

    cookie = handler.headers.get("Cookie")

    if not cookie:
        return None

    token=None
    partes=cookie.split(";")

    for parte in partes:

        if "session=" in parte:
            token=parte.strip().split("=")[1]
            break

    if not token:
        return None

    sesion=sesiones.get(token)

    if not sesion:
        return None

    usuario=sesion["usuario"]

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    SELECT programa
    FROM usuarios
    WHERE usuario=?
    """,(usuario,))

    resultado=cursor.fetchone()

    conexion.close()

    if resultado:
        return resultado[0]

    return None


# =================================================
# LISTAR
# =================================================

def listar_materias_usuario(handler):

    programa = obtener_programa_usuario(handler)
    if not programa:
        handler.send_response(401)
        handler.end_headers()
        handler.wfile.write(b"Sesion invalida")
        return

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    SELECT id,materia,continuidad,creditos
    FROM materias
    WHERE programa=?
    """,(programa,))

    datos = cursor.fetchall()

    conexion.close()

    lista=[]

    for m in datos:

        lista.append({

            "id":m[0],
            "materia":m[1],
            "continuidad":m[2],
            "creditos":m[3]

        })

    handler.send_response(200)
    handler.send_header("Content-type","application/json")
    handler.end_headers()

    handler.wfile.write(json.dumps(lista).encode())


# =================================================
# AGREGAR
# =================================================

def agregar_materia_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud = int(handler.headers['Content-Length'])
    datos = handler.rfile.read(longitud).decode()

    datos_json = json.loads(datos)

    materia = (datos_json["materia"] or "").strip()
    continuidad = (datos_json["continuidad"] or "").strip()
    creditos = datos_json["creditos"] or 0

    conexion = sqlite3.connect(config.DB_PATH)
    cursor = conexion.cursor()

    cursor.execute("""
    INSERT INTO materias
    (programa,materia,continuidad,creditos)
    VALUES (?,?,?,?)
    """,(programa,materia,continuidad,creditos))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Materia agregada")


# =================================================
# EDITAR
# =================================================

def editar_materia_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud=int(handler.headers['Content-Length'])
    datos=handler.rfile.read(longitud).decode()

    datos_json=json.loads(datos)
    
    materia = (datos_json["materia"] or "").strip()
    continuidad = (datos_json["continuidad"] or "").strip()
    creditos = datos_json["creditos"] or 0

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    UPDATE materias
    SET materia=?,continuidad=?,creditos=?
    WHERE id=? AND programa=?
    """,(materia,continuidad,creditos,datos_json["id"],programa))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Materia actualizada")


# =================================================
# ELIMINAR
# =================================================

def eliminar_materia_usuario(handler):

    programa = obtener_programa_usuario(handler)

    longitud=int(handler.headers['Content-Length'])
    datos=handler.rfile.read(longitud).decode()

    datos_json=json.loads(datos)

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    cursor.execute("""
    DELETE FROM materias
    WHERE id=? AND programa=?
    """,(

        datos_json["id"],
        programa

    ))

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(b"Materia eliminada")


# =================================================
# IMPORTAR EXCEL
# =================================================

def importar_materias_usuario(handler):

    import cgi
    from openpyxl import load_workbook
    import io

    programa = obtener_programa_usuario(handler)

    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            'REQUEST_METHOD':'POST',
            'CONTENT_TYPE':handler.headers['Content-Type']
        }
    )

    archivo=form['archivo']

    if not archivo.file:

        handler.send_response(400)
        handler.end_headers()
        handler.wfile.write(b"No se recibio archivo")
        return

    contenido=archivo.file.read()

    archivo_excel=io.BytesIO(contenido)

    wb=load_workbook(archivo_excel)

    hoja=wb.active

    conexion=sqlite3.connect(config.DB_PATH)
    cursor=conexion.cursor()

    agregados=0

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        materia,continuidad,creditos = fila

        materia = (materia or "").strip()
        continuidad = (continuidad or "").strip()
        creditos = creditos or 0

        if not materia:
            continue

        cursor.execute("""
        INSERT INTO materias
        (programa,materia,continuidad,creditos)
        VALUES (?,?,?,?)
        """,(

            programa,
            materia,
            continuidad,
            creditos

        ))

        agregados+=1

    conexion.commit()
    conexion.close()

    handler.send_response(200)
    handler.end_headers()

    handler.wfile.write(f"{agregados} materias importadas".encode())